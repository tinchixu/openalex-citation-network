"""
Pass 1: Match input papers to OpenAlex by DOI (primary) and title+year (fallback).

Accepts any CSV or xlsx file with columns: doi, title, year
  - doi   → required for primary matching
  - title + year → fallback when DOI is absent or mismatched

Fallback guards:
  - Titles shorter than 5 words are excluded (too generic)
  - title+year keys mapping to more than one paper are discarded (ambiguous)

Input : INPUT_FILE  (csv or xlsx)
Output: data/processed/matched_works.jsonl
"""

import gzip, re, csv
from pathlib import Path
from multiprocessing import Pool
import orjson
import openpyxl

# ── Configuration ─────────────────────────────────────────────────────────────

INPUT_FILE = Path("data/raw/sample_input.csv")    # ← change to your file
WORKS_DIR  = Path("/path/to/openalex/data/works") # ← path to OpenAlex snapshot
OUT_DIR    = Path("data/processed")
OUT_FILE   = OUT_DIR / "matched_works.jsonl"
N_WORKERS  = 6    # reduce to 2-3 for spinning HDDs; 6-8 for SSDs

# ── Worker globals ────────────────────────────────────────────────────────────

_DOI_SET        = None
_TITLE_YEAR_MAP = None

def init_worker(doi_set, title_year_map):
    global _DOI_SET, _TITLE_YEAR_MAP
    _DOI_SET        = doi_set
    _TITLE_YEAR_MAP = title_year_map

# ── Helpers ───────────────────────────────────────────────────────────────────

_RE_NORM = re.compile(r"[^\w\s]")          # precompiled: remove punctuation
_RE_SPC  = re.compile(r"\s+")             # precompiled: collapse whitespace

def norm_title(t):
    if not t:
        return ""
    return _RE_SPC.sub(" ", _RE_NORM.sub(" ", t.lower())).strip()

def _norm_doi(doi):
    return str(doi).strip().lower().removeprefix("https://doi.org/") if doi else ""

def load_input(path):
    """
    Load papers from a CSV or xlsx file.
    Recognised columns (case-insensitive): doi, title, year
    Returns: (doi_set, title_year_map)
    """
    path = Path(path)

    if path.suffix == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader   = csv.DictReader(f)
            raw_rows = [{k.lower().strip(): v for k, v in row.items()} for row in reader]
    elif path.suffix in (".xlsx", ".xls"):
        wb      = openpyxl.load_workbook(path, read_only=True)
        ws      = wb.active
        headers = [str(c.value).lower().strip() if c.value else ""
                   for c in next(ws.iter_rows(max_row=1))]
        raw_rows = [dict(zip(headers, row))
                    for row in ws.iter_rows(min_row=2, values_only=True)]
        wb.close()
    else:
        raise SystemExit(f"Unsupported input format: {path.suffix}  (use .csv or .xlsx)")

    doi_set   = set()
    ty_map    = {}
    ambiguous = set()

    for row in raw_rows:
        doi_norm = _norm_doi(row.get("doi"))
        if doi_norm:
            doi_set.add(doi_norm)

        title = row.get("title") or ""
        year  = row.get("year")
        if title and year:
            try:
                yr = int(float(str(year)))
            except (ValueError, TypeError):
                continue
            nt = norm_title(title)
            if len(nt.split()) >= 5:
                key = (nt, yr)
                if key in ty_map:
                    ambiguous.add(key)
                else:
                    ty_map[key] = doi_norm

    for key in ambiguous:
        del ty_map[key]

    return doi_set, ty_map

def scan_file(gz_path):
    results = []
    try:
        with gzip.open(gz_path, "rb") as f:
            for line in f:
                if not line or line == b"\n":
                    continue
                w      = orjson.loads(line)
                oa_doi = _norm_doi(w.get("doi"))

                # Primary: DOI match
                if oa_doi and oa_doi in _DOI_SET:
                    results.append({
                        "id":               w["id"],
                        "doi":              oa_doi,
                        "match_method":     "doi",
                        "title":            w.get("display_name"),
                        "year":             w.get("publication_year"),
                        "referenced_works": w.get("referenced_works", []),
                    })
                    continue

                # Fallback: title + year (check year first — cheaper)
                yr = w.get("publication_year")
                if not yr:
                    continue
                nt  = norm_title(w.get("display_name"))
                key = (nt, yr)
                if nt and key in _TITLE_YEAR_MAP:
                    results.append({
                        "id":               w["id"],
                        "doi":              oa_doi or _TITLE_YEAR_MAP[key],
                        "wos_doi":          _TITLE_YEAR_MAP[key],
                        "match_method":     "title_year",
                        "title":            w.get("display_name"),
                        "year":             yr,
                        "referenced_works": w.get("referenced_works", []),
                    })
    except Exception as e:
        print(f"  [error] {gz_path}: {e}")
    return results

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Loading input: {INPUT_FILE}")
    doi_set, title_year_map = load_input(INPUT_FILE)
    print(f"  {len(doi_set):,} DOIs (primary matching)")
    print(f"  {len(title_year_map):,} unambiguous title+year keys (fallback)")

    files = sorted(WORKS_DIR.rglob("*.gz"))
    if not files:
        raise SystemExit(f"No .gz files found in WORKS_DIR: {WORKS_DIR}\n"
                         "Check the path in the Configuration section.")
    print(f"  {len(files):,} snapshot files to scan")
    print(f"  Writing to {OUT_FILE}\n")

    n_doi, n_title = 0, 0
    seen = set()

    # 8 MB write buffer reduces syscall overhead
    with open(OUT_FILE, "wb", buffering=8 * 1024 * 1024) as out, \
         Pool(N_WORKERS, initializer=init_worker,
              initargs=(doi_set, title_year_map)) as pool:
        for i, batch in enumerate(
            pool.imap_unordered(scan_file, map(str, files), chunksize=32), 1
        ):
            for r in batch:
                if r["id"] in seen:
                    continue
                seen.add(r["id"])
                out.write(orjson.dumps(r) + b"\n")
                if r["match_method"] == "doi":
                    n_doi += 1
                else:
                    n_title += 1
            if i % 200 == 0:
                print(f"  {i}/{len(files)} files | doi={n_doi:,}  title+year={n_title:,}")

    print(f"\nDone:")
    print(f"  Matched by DOI:        {n_doi:,}")
    print(f"  Matched by title+year: {n_title:,}")
    print(f"  Total:                 {n_doi + n_title:,}")
    print(f"  Output → {OUT_FILE}")

if __name__ == "__main__":
    main()
