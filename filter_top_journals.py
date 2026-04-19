"""
Filter the WoS master xlsx to papers published in top economics and finance journals.

Target journals:
  Economics  : AER, QJE, Journal of Political Economy
               AEJ Applied, AEJ Policy, AEJ Macro, AEJ Micro
  Finance    : Journal of Finance, Review of Financial Studies,
               Journal of Financial Economics
  Econometrics: Journal of Econometrics

Input : data/raw/wos_docs_readable.xlsx
Output: data/raw/wos_top_journals.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import Counter

IN_FILE  = Path("data/raw/wos_top_journals.xlsx")   # already journal-filtered
OUT_FILE = Path("data/raw/wos_top_journals.xlsx")  # overwrite in place

TOP_JOURNALS = {
    # Economics
    "AMERICAN ECONOMIC REVIEW",
    "QUARTERLY JOURNAL OF ECONOMICS",
    "JOURNAL OF POLITICAL ECONOMY",
    "AMERICAN ECONOMIC JOURNAL-APPLIED ECONOMICS",
    "AMERICAN ECONOMIC JOURNAL-ECONOMIC POLICY",
    "AMERICAN ECONOMIC JOURNAL-MACROECONOMICS",
    "AMERICAN ECONOMIC JOURNAL-MICROECONOMICS",
    # Finance
    "JOURNAL OF FINANCE",
    "REVIEW OF FINANCIAL STUDIES",
    "JOURNAL OF FINANCIAL ECONOMICS",
    # Econometrics
    "JOURNAL OF ECONOMETRICS",
}

# ── Read ──────────────────────────────────────────────────────────────────────

print(f"Reading {IN_FILE} ...")
wb_in = openpyxl.load_workbook(IN_FILE, read_only=True)
ws_in = wb_in.active

headers  = [c.value for c in next(ws_in.iter_rows(max_row=1))]
src_idx  = headers.index("source_title")
type_idx = headers.index("source_types")

rows = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    if (row[src_idx] and str(row[src_idx]).strip().upper() in TOP_JOURNALS
            and row[type_idx] == "Article"):
        rows.append(row)

wb_in.close()
print(f"  {len(rows):,} papers matched out of original dataset")

# ── Journal breakdown ─────────────────────────────────────────────────────────

counts = Counter(str(r[src_idx]).strip() for r in rows)
print("\nPapers per journal:")
for journal, n in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {n:>6,}  {journal}")

# ── Write ─────────────────────────────────────────────────────────────────────

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Top Journals"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")

# Write header
for col, name in enumerate(headers, 1):
    cell = ws_out.cell(row=1, column=col, value=name)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center")

# Write data rows
for row_idx, row in enumerate(rows, 2):
    for col, val in enumerate(row, 1):
        ws_out.cell(row=row_idx, column=col, value=val)

# Auto-fit column widths (capped at 50)
for col, name in enumerate(headers, 1):
    max_len = max(
        (len(str(r[col - 1]) if r[col - 1] is not None else "") for r in rows),
        default=0,
    )
    ws_out.column_dimensions[get_column_letter(col)].width = min(max(max_len, len(str(name))) + 2, 50)

ws_out.freeze_panes = "A2"

wb_out.save(OUT_FILE)
print(f"\nSaved → {OUT_FILE}  ({len(rows):,} rows × {len(headers)} columns)")
