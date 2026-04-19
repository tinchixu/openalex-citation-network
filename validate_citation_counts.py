"""
Validation 1: WoS times_cited vs OpenAlex incoming citation count.

Restricted to WoS document type "Article" (original research papers only).
Excludes: Review, Editorial, Letter, Correction, Abstract, retracted, etc.

WoS and OpenAlex differ by design (WoS = curated journals only, OpenAlex = broader),
so we expect OpenAlex >= WoS on average. We validate on correlation and flag outliers.

Inputs:
  data/raw/wos_docs_readable.xlsx
  data/processed/matched_works.jsonl      (openalex_id → doi)
  data/processed/citations_edges.csv      (source → target openalex_id)

Output:
  data/processed/validation_citation_counts.csv
  data/processed/validation_citation_counts.png
"""

import json, csv
from pathlib import Path
from collections import Counter

import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

RAW_DIR   = Path("data/raw")
OUT_DIR   = Path("data/processed")
XLSX      = RAW_DIR / "wos_docs_readable.xlsx"
MATCHED   = OUT_DIR / "matched_works.jsonl"
CIT_EDGES = OUT_DIR / "citations_edges.csv"
OUT_CSV   = OUT_DIR / "validation_citation_counts.csv"
OUT_PNG   = OUT_DIR / "validation_citation_counts.png"

# ── 1. Load WoS doi → times_cited  (Articles only) ───────────────────────────

# Included WoS types: anything starting with "Article" except retracted.
# Excludes: Review, Editorial material, Letter, Correction, Abstract, etc.
ARTICLE_TYPES = {"Article", "Article; Early Access", "Article; Meeting", "Article; Book"}

print("Loading WoS citation counts (Articles only)...")
wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb.active
headers  = [c.value for c in next(ws.iter_rows(max_row=1))]
doi_idx  = headers.index("doi")
cite_idx = headers.index("times_cited")
type_idx = headers.index("types")

wos = {}   # doi (normalised) → times_cited
skipped = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    doi  = row[doi_idx]
    tc   = row[cite_idx]
    wtype = row[type_idx]
    if wtype not in ARTICLE_TYPES:
        skipped[wtype] += 1
        continue
    if doi and tc is not None:
        wos[str(doi).strip().lower().removeprefix("https://doi.org/")] = int(tc)
wb.close()
print(f"  {len(wos):,} research articles with citation counts")
print(f"  Excluded types:")
for t, n in skipped.most_common():
    print(f"    {n:>8,}  {t}")

# ── 2. Load matched_works: openalex_id → doi ─────────────────────────────────

print("Loading matched works...")
id_to_doi = {}
with open(MATCHED) as f:
    for line in f:
        w = json.loads(line)
        id_to_doi[w["id"]] = w["doi"]
print(f"  {len(id_to_doi):,} matched papers")

# ── 3. Count incoming citations per openalex_id from edge list ────────────────

print("Counting OpenAlex incoming citations...")
oa_counts = Counter()
with open(CIT_EDGES) as f:
    reader = csv.reader(f)
    next(reader)   # skip header
    for source, target in reader:
        if target in id_to_doi:
            oa_counts[target] += 1
print(f"  {len(oa_counts):,} papers received at least one citation in OpenAlex")

# ── 4. Merge and compare ──────────────────────────────────────────────────────

print("Merging...")
rows = []
for oa_id, doi in id_to_doi.items():
    wos_tc = wos.get(doi)
    oa_tc  = oa_counts.get(oa_id, 0)
    if wos_tc is not None:
        rows.append({"doi": doi, "openalex_id": oa_id,
                     "wos_cited": wos_tc, "oa_cited": oa_tc,
                     "difference": oa_tc - wos_tc})

print(f"  {len(rows):,} papers in both datasets")

wos_arr = np.array([r["wos_cited"] for r in rows])
oa_arr  = np.array([r["oa_cited"]  for r in rows])
diff    = oa_arr - wos_arr

print(f"\n── Summary ────────────────────────────────────")
print(f"  Median WoS citations:      {np.median(wos_arr):.1f}")
print(f"  Median OpenAlex citations: {np.median(oa_arr):.1f}")
print(f"  Median difference (OA-WoS): {np.median(diff):.1f}")
print(f"  Papers where OA > WoS:  {(diff > 0).sum():,} ({(diff > 0).mean()*100:.1f}%)")
print(f"  Papers where OA = WoS:  {(diff == 0).sum():,} ({(diff == 0).mean()*100:.1f}%)")
print(f"  Papers where OA < WoS:  {(diff < 0).sum():,} ({(diff < 0).mean()*100:.1f}%)")

# Correlation on log scale (add 1 to handle zeros)
log_wos = np.log1p(wos_arr)
log_oa  = np.log1p(oa_arr)
corr = np.corrcoef(log_wos, log_oa)[0, 1]
print(f"  Pearson r (log scale):  {corr:.4f}")

# ── 5. Save CSV ───────────────────────────────────────────────────────────────

rows.sort(key=lambda r: r["difference"])   # worst undercount first
with open(OUT_CSV, "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["doi","openalex_id","wos_cited","oa_cited","difference"])
    writer.writeheader()
    writer.writerows(rows)
print(f"\nCSV saved → {OUT_CSV}")

# ── 6. Plot ───────────────────────────────────────────────────────────────────

fig, axes = plt.subplots(1, 2, figsize=(12, 5))
fig.suptitle("WoS vs OpenAlex Citation Count Validation  (Articles only)", fontsize=13, fontweight="bold")

# Scatter: log-log
ax = axes[0]
ax.scatter(wos_arr + 1, oa_arr + 1, s=1, alpha=0.15, color="steelblue", rasterized=True)
lim = max(wos_arr.max(), oa_arr.max()) * 1.5
ax.set_xlim(1, lim); ax.set_ylim(1, lim)
ax.set_xscale("log"); ax.set_yscale("log")
ax.plot([1, lim], [1, lim], "r--", lw=1, label="y = x")
ax.set_xlabel("WoS times_cited + 1")
ax.set_ylabel("OpenAlex incoming citations + 1")
ax.set_title(f"Log-log scatter  (r = {corr:.3f})")
ax.legend(fontsize=9)

# Histogram of differences (capped for readability)
ax = axes[1]
cap = np.percentile(np.abs(diff), 99)
clipped = np.clip(diff, -cap, cap)
ax.hist(clipped, bins=80, color="steelblue", edgecolor="none", alpha=0.8)
ax.axvline(0, color="red", lw=1.2, linestyle="--", label="no difference")
ax.axvline(np.median(diff), color="orange", lw=1.2, linestyle="-", label=f"median = {np.median(diff):.0f}")
ax.set_xlabel("OpenAlex − WoS (citations, 99th pct clipped)")
ax.set_ylabel("Number of papers")
ax.set_title("Distribution of difference")
ax.legend(fontsize=9)
ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"{int(x):,}"))

plt.tight_layout()
plt.savefig(OUT_PNG, dpi=150, bbox_inches="tight")
print(f"Plot saved → {OUT_PNG}")
plt.show()
