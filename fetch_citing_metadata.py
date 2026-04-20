"""
For a given DOI, export all papers that cite it — with full metadata —
from the locally processed citing_works.jsonl. No API calls needed.

Once exported, filter/sort/analyse in Excel as usual.

Inputs:
  data/processed/matched_works.jsonl   — resolves DOI → OpenAlex ID
  data/processed/citing_works.jsonl    — full metadata for all citing papers

Output:
  data/processed/citing_{slug}.xlsx    — one sheet, all citing papers
"""

import re
from pathlib import Path
from collections import Counter
import orjson
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuration — edit this ─────────────────────────────────────────────────

TARGET_DOI = "10.1086/260062"   # ← change to any DOI in your dataset

OUT_DIR = Path("data/processed")

# ── Helpers ───────────────────────────────────────────────────────────────────

def doi_slug(doi):
    return re.sub(r"[^\w]", "_", doi)

def find_openalex_id(doi):
    with open(OUT_DIR / "matched_works.jsonl", "rb") as f:
        for line in f:
            w = orjson.loads(line)
            if w.get("doi") == doi or w.get("input_doi") == doi:
                return w["id"], w.get("title"), w.get("year")
    return None, None, None

def write_sheet(ws, data, fields):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for col, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, record in enumerate(data, 2):
        for col, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col, value=record.get(field))
    for col, field in enumerate(fields, 1):
        max_len = max((len(str(r.get(field) or "")) for r in data), default=0)
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len, len(field)) + 2, 60)
    ws.freeze_panes = "A2"

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    slug = doi_slug(TARGET_DOI)

    print(f"Looking up: {TARGET_DOI}")
    oa_id, title, year = find_openalex_id(TARGET_DOI)
    if not oa_id:
        raise SystemExit(f"DOI not found in matched_works.jsonl: {TARGET_DOI}")
    print(f"  {title} ({year})")
    print(f"  {oa_id}")

    print("\nReading citing_works.jsonl...")
    rows = []
    with open(OUT_DIR / "citing_works.jsonl", "rb") as f:
        for line in f:
            w = orjson.loads(line)
            if oa_id in w.get("cites_our", []):
                rows.append(w)
    print(f"  {len(rows):,} citing papers found")

    fields = ["id", "doi", "title", "year", "type", "journal", "issn_l",
              "authors", "institutions", "cited_by_count", "oa_status"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Citing Papers"
    write_sheet(ws, rows, fields)

    out_path = OUT_DIR / f"citing_{slug}.xlsx"
    wb.save(out_path)
    print(f"\nOutput → {out_path}  ({len(rows):,} rows)")

    # Top 20 journals among citers
    counts = Counter(r.get("journal") for r in rows if r.get("journal"))
    print("\nTop 20 journals citing this paper:")
    for journal, n in counts.most_common(20):
        print(f"  {n:>6}  {journal}")

if __name__ == "__main__":
    main()
